from __future__ import annotations

import csv
import json
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs"
THUMB_DIR = OUTPUT_DIR / "excel_thumbnails"
OUT_FILE = OUTPUT_DIR / "gom_experiment_benchmark_report.xlsx"

DATASET1_ROOT = ROOT / "dataset" / "video_experiment_100"
DATASET2_ROOT = ROOT / "dataset" / "ai_generated_collection_100"
DATASET1_MANIFEST = DATASET1_ROOT / "manifest.csv"
DATASET2_MANIFEST = DATASET2_ROOT / "manifest.csv"
DATASET1_RESULTS = ROOT / "gom-ai" / "experiment_results" / "dataset1_video_lens5" / "detailed_results.json"
DATASET1_SUMMARY = ROOT / "gom-ai" / "experiment_results" / "dataset1_video_lens5" / "summary.json"
DATASET2_RESULTS = ROOT / "gom-ai" / "experiment_results" / "dataset2_ai_lens5" / "detailed_results.json"
DATASET2_SUMMARY = ROOT / "gom-ai" / "experiment_results" / "dataset2_ai_lens5" / "summary.json"

METHODS = ["gemini", "chatgpt", "grok", "acis"]
METHOD_LABELS = {
    "gemini": "Gemini",
    "chatgpt": "ChatGPT",
    "grok": "Groq/Llama",
    "acis": "ACIS",
}


def read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def clean_sheet_title(title: str) -> str:
    return title[:31]


def set_title(ws, title: str, subtitle: str | None = None, last_col: int = 8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, title)
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center")
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        sub = ws.cell(2, 1, subtitle)
        sub.font = Font(italic=True, color="666666")
        sub.alignment = Alignment(wrap_text=True)


def style_header(ws, row: int):
    fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(bottom=Side(style="thin", color="9EADBD"))
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width: int = 42):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        best = 10
        for cell in ws[letter]:
            if cell.value is None:
                continue
            best = max(best, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = best


def thumb_for(image_path: Path, name: str) -> Path | None:
    if not image_path.exists():
        return None
    THUMB_DIR.mkdir(parents=True, exist_ok=True)
    target = THUMB_DIR / f"{name}.jpg"
    with PILImage.open(image_path) as img:
        img = img.convert("RGB")
        img.thumbnail((96, 96), PILImage.Resampling.LANCZOS)
        canvas = PILImage.new("RGB", (96, 96), "white")
        x = (96 - img.width) // 2
        y = (96 - img.height) // 2
        canvas.paste(img, (x, y))
        canvas.save(target, "JPEG", quality=82)
    return target


def add_thumb(ws, row: int, col: int, image_path: Path | None, name: str):
    if not image_path:
        return
    thumb = thumb_for(image_path, name)
    if not thumb:
        return
    img = ExcelImage(str(thumb))
    img.width = 72
    img.height = 72
    ws.add_image(img, f"{get_column_letter(col)}{row}")
    ws.row_dimensions[row].height = 58


def method_complete(row: dict) -> bool:
    if row.get("run_error"):
        return False
    methods = row.get("methods", {})
    return all(not methods.get(method, {}).get("error") for method in METHODS)


def build_summary_records(summary: dict) -> list[list]:
    rows = []
    for method in METHODS:
        metrics = summary.get("methods", {}).get(method)
        if not metrics:
            rows.append([METHOD_LABELS[method], "Pending", "", "", "", "", "", "", ""])
            continue
        rows.append(
            [
                METHOD_LABELS[method],
                metrics.get("total"),
                metrics.get("correct"),
                metrics.get("errors"),
                metrics.get("accuracy"),
                metrics.get("macro_precision"),
                metrics.get("macro_recall"),
                metrics.get("macro_f1_score"),
                metrics.get("average_latency_s"),
                metrics.get("hallucination_rate"),
            ]
        )
    return rows


def add_readme(wb: Workbook):
    ws = wb.active
    ws.title = "README"
    set_title(ws, "Ceramic AI Experiment Workbook", "Two 100-sample experiments: real/video dataset and AI-generated dataset.", 8)
    rows = [
        ["Field", "Value"],
        ["Experiment 1", "100 real/video ceramic images. Benchmark results currently available."],
        ["Experiment 2", "100 AI-generated ceramic images. Benchmark results currently available."],
        ["Compared systems", "Gemini, ChatGPT, Groq/Llama, ACIS multi-agent workflow"],
        ["Accuracy", "N_correct / N_total * 100"],
        ["Precision", "TP / (TP + FP), macro averaged across 10 ceramic classes"],
        ["Recall", "TP / (TP + FN), macro averaged across 10 ceramic classes"],
        ["F1-score", "2 * Precision * Recall / (Precision + Recall)"],
        ["Hallucination rate", "Predictions outside the supported 10 labels / N_total * 100"],
        ["Runs + average /100", "For each experiment and method, the workbook records evaluated runs, correct count, error count, hallucination count, and per-100 averages."],
        ["Important note", "Rows with API/vision errors are retained and flagged; rerun them before final submission if a fully clean 100/100 benchmark is required."],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 4)
    auto_width(ws, 80)
    ws.freeze_panes = "A5"


def add_dataset1_summary(wb: Workbook, summary: dict, results: list[dict]):
    ws = wb.create_sheet("Dataset1_Summary")
    set_title(ws, "Dataset 1 Summary", "Real/video image dataset benchmark. Metrics are from detailed_results.json.", 10)
    complete = sum(method_complete(row) for row in results)
    ws.append([])
    ws.append(["Dataset size", len(results), "Complete rows", complete, "Rows with errors", len(results) - complete])
    ws.append([])
    headers = ["Method", "N total", "N correct", "Errors", "Accuracy %", "Precision %", "Recall %", "F1 %", "Avg latency s", "Hallucination %"]
    ws.append(headers)
    for row in build_summary_records(summary):
        ws.append(row)
    style_header(ws, 5)
    for row in ws.iter_rows(min_row=6, max_row=9, min_col=5, max_col=10):
        for cell in row:
            cell.number_format = "0.00"
    auto_width(ws)
    ws.freeze_panes = "A6"


def add_dataset2_summary(wb: Workbook, summary: dict, results: list[dict]):
    ws = wb.create_sheet("Dataset2_Summary")
    set_title(ws, "Dataset 2 Summary", "AI-generated image dataset benchmark. Metrics are from detailed_results.json.", 10)
    complete = sum(method_complete(row) for row in results)
    ws.append([])
    ws.append(["Dataset size", len(results), "Complete rows", complete, "Rows with errors", len(results) - complete])
    ws.append([])
    ws.append(["Method", "N total", "N correct", "Errors", "Accuracy %", "Precision %", "Recall %", "F1 %", "Avg latency s", "Hallucination %"])
    for row in build_summary_records(summary):
        ws.append(row)
    style_header(ws, 5)
    for row in ws.iter_rows(min_row=6, max_row=9, min_col=5, max_col=10):
        for cell in row:
            cell.number_format = "0.00"
    auto_width(ws)
    ws.freeze_panes = "A6"


def add_final_comparison(wb: Workbook, dataset1_summary: dict, dataset2_summary: dict):
    ws = wb.create_sheet("Final_Comparison", 1)
    set_title(ws, "Final Comparison", "Average metrics per 100 samples for the two experiments.", 11)
    ws.append([])
    ws.append(["Experiment", "Method", "N total", "N correct", "Errors", "Accuracy %", "Precision %", "Recall %", "F1 %", "Avg latency s", "Hallucination %"])
    for row in build_summary_records(dataset1_summary):
        ws.append(["Dataset 1 - Real/video"] + row)
    for row in build_summary_records(dataset2_summary):
        ws.append(["Dataset 2 - AI-generated"] + row)
    style_header(ws, 4)
    for row in ws.iter_rows(min_row=5, max_row=12, min_col=6, max_col=11):
        for cell in row:
            cell.number_format = "0.00"
    auto_width(ws)
    ws.freeze_panes = "A5"


def add_runs_average(wb: Workbook, dataset_summaries: list[tuple[str, dict]]):
    ws = wb.create_sheet("Runs_Average_100", 2)
    set_title(ws, "Runs + Average /100", "Count-based and average metrics for the two 100-sample experiments.", 14)
    ws.append([])
    headers = [
        "Experiment", "Method", "Runs evaluated", "Correct count", "Correct avg /100",
        "Error count", "Error avg /100", "Hallucination count", "Hallucination avg /100",
        "Accuracy %", "Precision %", "Recall %", "F1 %", "Avg latency s",
    ]
    ws.append(headers)
    style_header(ws, 4)
    for experiment, summary in dataset_summaries:
        for method in METHODS:
            metrics = summary.get("methods", {}).get(method, {})
            total = metrics.get("total") or 0
            correct = metrics.get("correct") or 0
            errors = metrics.get("errors") or 0
            hallucinated = metrics.get("hallucinated") or 0
            per_100 = 100 / total if total else 0
            ws.append([
                experiment,
                METHOD_LABELS[method],
                total,
                correct,
                round(correct * per_100, 2) if total else 0,
                errors,
                round(errors * per_100, 2) if total else 0,
                hallucinated,
                round(hallucinated * per_100, 2) if total else 0,
                metrics.get("accuracy"),
                metrics.get("macro_precision"),
                metrics.get("macro_recall"),
                metrics.get("macro_f1_score"),
                metrics.get("average_latency_s"),
            ])

    ws.append([])
    ws.append(["Combined average across 2 experiments", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    combined_header_row = ws.max_row + 1
    ws.append(["Method", "Avg correct /100", "Avg errors /100", "Avg hallucination /100", "Avg accuracy %", "Avg precision %", "Avg recall %", "Avg F1 %", "Avg latency s"])
    style_header(ws, combined_header_row)
    for method in METHODS:
        values = []
        for _, summary in dataset_summaries:
            metrics = summary.get("methods", {}).get(method, {})
            total = metrics.get("total") or 0
            per_100 = 100 / total if total else 0
            values.append({
                "correct": (metrics.get("correct") or 0) * per_100 if total else 0,
                "errors": (metrics.get("errors") or 0) * per_100 if total else 0,
                "hallucinated": (metrics.get("hallucinated") or 0) * per_100 if total else 0,
                "accuracy": metrics.get("accuracy") or 0,
                "precision": metrics.get("macro_precision") or 0,
                "recall": metrics.get("macro_recall") or 0,
                "f1": metrics.get("macro_f1_score") or 0,
                "latency": metrics.get("average_latency_s") or 0,
            })
        ws.append([
            METHOD_LABELS[method],
            round(sum(v["correct"] for v in values) / len(values), 2),
            round(sum(v["errors"] for v in values) / len(values), 2),
            round(sum(v["hallucinated"] for v in values) / len(values), 2),
            round(sum(v["accuracy"] for v in values) / len(values), 2),
            round(sum(v["precision"] for v in values) / len(values), 2),
            round(sum(v["recall"] for v in values) / len(values), 2),
            round(sum(v["f1"] for v in values) / len(values), 2),
            round(sum(v["latency"] for v in values) / len(values), 3),
        ])

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=14):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00" if isinstance(cell.value, float) else "0"
    auto_width(ws, 48)
    ws.freeze_panes = "A5"


def add_dataset1_results(wb: Workbook, manifest_rows: list[dict], results: list[dict]):
    ws = wb.create_sheet("Dataset1_Results")
    set_title(ws, "Dataset 1 Results", "Each row includes image, source, ground truth, predictions and benchmark flags.", 22)
    by_id = {int(row["id"]): row for row in results}
    headers = [
        "Index", "Image", "Filename", "Source URL", "Timestamp", "Country", "Ground truth",
        "Complete?", "Run error",
        "Gemini pred", "Gemini correct", "Gemini latency", "Gemini error",
        "ChatGPT pred", "ChatGPT correct", "ChatGPT latency", "ChatGPT error",
        "Groq/Llama pred", "Groq correct", "Groq latency", "Groq error",
        "ACIS pred", "ACIS correct", "ACIS latency", "ACIS error",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws, 4)
    row_num = 5
    for m in manifest_rows:
        idx = int(m["index"])
        result = by_id.get(idx, {})
        methods = result.get("methods", {})
        values = [
            idx,
            "",
            m.get("filename"),
            m.get("source_url"),
            m.get("timestamp_hhmmss"),
            m.get("country"),
            result.get("ground_truth") or m.get("ceramic_tradition"),
            "Yes" if result and method_complete(result) else "No",
            result.get("run_error"),
        ]
        for method in METHODS:
            r = methods.get(method, {})
            values.extend([
                r.get("predicted_label") or r.get("raw_prediction"),
                r.get("is_correct"),
                r.get("latency_s"),
                r.get("error"),
            ])
        ws.append(values)
        image_path = DATASET1_ROOT / m.get("filename", "")
        add_thumb(ws, row_num, 2, image_path, f"d1_{idx:03d}")
        row_num += 1
    ws.auto_filter.ref = f"A4:Y{ws.max_row}"
    ws.freeze_panes = "A5"
    ws.column_dimensions["B"].width = 14
    auto_width(ws, 55)
    ws.column_dimensions["B"].width = 14


def add_dataset1_metadata(wb: Workbook, manifest_rows: list[dict]):
    ws = wb.create_sheet("Dataset1_Metadata")
    set_title(ws, "Dataset 1 Metadata", "Source information for 100 real/video samples.", 15)
    headers = [
        "Index", "Image", "Filename", "Tradition", "Region", "Country", "Video title",
        "Channel", "Source URL", "Timestamp", "Width", "Height", "License", "Accessed date",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws, 4)
    row_num = 5
    for m in manifest_rows:
        idx = int(m["index"])
        ws.append([
            idx, "", m.get("filename"), m.get("ceramic_tradition"), m.get("region"),
            m.get("country"), m.get("video_title"), m.get("channel"),
            m.get("source_url"), m.get("timestamp_hhmmss"), m.get("width"),
            m.get("height"), m.get("license"), m.get("accessed_date"),
        ])
        add_thumb(ws, row_num, 2, DATASET1_ROOT / m.get("filename", ""), f"d1_meta_{idx:03d}")
        row_num += 1
    ws.auto_filter.ref = f"A4:N{ws.max_row}"
    ws.freeze_panes = "A5"
    auto_width(ws, 55)
    ws.column_dimensions["B"].width = 14


def add_dataset2_metadata(wb: Workbook, manifest_rows: list[dict]):
    ws = wb.create_sheet("Dataset2_Metadata")
    set_title(ws, "Dataset 2 Metadata", "AI-generated 100-image dataset from ceramic collection references.", 16)
    headers = [
        "Index", "Image", "Filename", "Shuffled filename", "Tradition", "Country", "Region",
        "Object type", "Reference image", "Reference URL", "Generation method", "Created date", "Synthetic",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws, 4)
    row_num = 5
    for m in manifest_rows:
        idx = int(m["index"])
        ws.append([
            idx, "", m.get("filename"), m.get("shuffled_filename"), m.get("tradition"),
            m.get("country"), m.get("region"), m.get("object_type"), m.get("reference_image"),
            m.get("reference_url"), m.get("generation_method"), m.get("created_date"),
            m.get("synthetic"),
        ])
        add_thumb(ws, row_num, 2, DATASET2_ROOT / m.get("filename", ""), f"d2_meta_{idx:03d}")
        row_num += 1
    ws.auto_filter.ref = f"A4:M{ws.max_row}"
    ws.freeze_panes = "A5"
    auto_width(ws, 55)
    ws.column_dimensions["B"].width = 14


def add_dataset2_results(wb: Workbook, manifest_rows: list[dict], results: list[dict]):
    ws = wb.create_sheet("Dataset2_Results")
    set_title(ws, "Dataset 2 Results", "Each row includes image, source, ground truth, predictions and benchmark flags.", 25)
    by_id = {int(row["id"]): row for row in results}
    headers = [
        "Index", "Image", "Filename", "Reference URL", "Generation method", "Object type", "Ground truth",
        "Complete?", "Run error",
        "Gemini pred", "Gemini correct", "Gemini latency", "Gemini error",
        "ChatGPT pred", "ChatGPT correct", "ChatGPT latency", "ChatGPT error",
        "Groq/Llama pred", "Groq correct", "Groq latency", "Groq error",
        "ACIS pred", "ACIS correct", "ACIS latency", "ACIS error",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws, 4)
    row_num = 5
    for m in manifest_rows:
        idx = int(m["index"])
        result = by_id.get(idx, {})
        methods = result.get("methods", {})
        values = [
            idx, "", m.get("filename"), m.get("reference_url"), m.get("generation_method"),
            m.get("object_type"), result.get("ground_truth") or m.get("tradition"),
            "Yes" if result and method_complete(result) else "No",
            result.get("run_error"),
        ]
        for method in METHODS:
            r = methods.get(method, {})
            values.extend([
                r.get("predicted_label") or r.get("raw_prediction"),
                r.get("is_correct"),
                r.get("latency_s"),
                r.get("error"),
            ])
        ws.append([
            *values,
        ])
        add_thumb(ws, row_num, 2, DATASET2_ROOT / m.get("filename", ""), f"d2_res_{idx:03d}")
        row_num += 1
    ws.auto_filter.ref = f"A4:Y{ws.max_row}"
    ws.freeze_panes = "A5"
    auto_width(ws, 55)
    ws.column_dimensions["B"].width = 14


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if THUMB_DIR.exists():
        shutil.rmtree(THUMB_DIR)
    d1_manifest = read_csv(DATASET1_MANIFEST)[:5]
    d2_manifest = read_csv(DATASET2_MANIFEST)[:5]
    d1_results = read_json(DATASET1_RESULTS, [])
    d1_summary = read_json(DATASET1_SUMMARY, {})
    d2_results = read_json(DATASET2_RESULTS, [])
    d2_summary = read_json(DATASET2_SUMMARY, {})

    wb = Workbook()
    add_readme(wb)
    add_final_comparison(wb, d1_summary, d2_summary)
    add_runs_average(wb, [
        ("Dataset 1 - Real/video", d1_summary),
        ("Dataset 2 - AI-generated", d2_summary),
    ])
    add_dataset1_summary(wb, d1_summary, d1_results)
    add_dataset1_metadata(wb, d1_manifest)
    add_dataset1_results(wb, d1_manifest, d1_results)
    add_dataset2_summary(wb, d2_summary, d2_results)
    add_dataset2_metadata(wb, d2_manifest)
    add_dataset2_results(wb, d2_manifest, d2_results)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.sheet_view.showGridLines = False

    try:
        wb.save(OUT_FILE)
        print(OUT_FILE)
    except PermissionError:
        alt = OUT_FILE.with_name("gom_experiment_benchmark_report_v2.xlsx")
        wb.save(alt)
        print(f"Original file locked. Saved to: {alt}")


if __name__ == "__main__":
    main()
