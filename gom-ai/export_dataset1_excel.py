"""Export real dataset 1 benchmark results comparing the ACIS System and 3 individual baseline models."""
import json
import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Load dataset 1 detailed results
results_path = "experiment_results/dataset1_video_lens5/detailed_results.json"
summary_path = "experiment_results/dataset1_video_lens5/summary.json"

if not os.path.exists(results_path) or not os.path.exists(summary_path):
    print("Error: Dataset 1 results or summary JSON files not found!")
    exit(1)

with open(results_path, "r", encoding="utf-8") as f:
    detailed_data = json.load(f)

with open(summary_path, "r", encoding="utf-8") as f:
    summary_data = json.load(f)

wb = openpyxl.Workbook()

# Define Color Palette
navy_dark = "1B365D"
accent_ice = "E8F1F5"
gray_border = "D3D3D3"
white = "FFFFFF"
acis_highlight = "E2F0D9"

font_title = Font(name="Arial", size=15, bold=True, color=navy_dark)
font_section = Font(name="Arial", size=11, bold=True, color=navy_dark)
font_header = Font(name="Arial", size=9, bold=True, color=white)
font_data = Font(name="Arial", size=9)
font_bold_data = Font(name="Arial", size=9, bold=True)

fill_header = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
fill_zebra = PatternFill(start_color=accent_ice, end_color=accent_ice, fill_type="solid")
fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_acis = PatternFill(start_color=acis_highlight, end_color=acis_highlight, fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

thin_side = Side(border_style="thin", color=gray_border)
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# ==========================================
# SHEET 1: SUMMARY METRICS (ACIS vs 3 Baselines)
# ==========================================
ws_summary = wb.active
ws_summary.title = "Báo cáo Tổng hợp"

# Title
ws_summary["A1"] = "BÁO CÁO THỰC NGHIỆM ĐỘ CHÍNH XÁC HỆ THỐNG ACIS VỚI CÁC AI ĐƠN LẺ"
ws_summary["A1"].font = font_title
ws_summary["A2"] = "Tập dữ liệu: dataset1_video_lens5 — 100 mẫu ảnh gốm sứ thực tế (trích xuất từ video)"
ws_summary["A2"].font = Font(name="Arial", size=10, italic=True)

ws_summary["A4"] = "Bảng so sánh hiệu năng tổng thể (100 mẫu):"
ws_summary["A4"].font = font_section

# Headers
headers_summary = [
    "Mô hình / Phương pháp",
    "Mô hình LLM Nền tảng",
    "Accuracy (%)",
    "Precision (%)",
    "Recall (%)",
    "F1-score (%)",
    "Thời gian TB (s)",
    "Hallucination (%)"
]
for col_idx, text in enumerate(headers_summary, 1):
    cell = ws_summary.cell(row=5, column=col_idx, value=text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

# Read from summary.json with CORRECT keys
methods_info = [
    ("chatgpt", "ChatGPT (Đơn lẻ)", "gpt-4o-mini"),
    ("gemini", "Gemini (Đơn lẻ)", "gemini-1.5-flash-8b"),
    ("grok", "Groq (Đơn lẻ)", "llama-3.1-8b-instant"),
    ("acis", "Hệ thống ACIS (Đa tác tử đồng thuận)", "Đa mô hình (ACIS Pipeline)"),
]

row_idx = 6
for method_key, display_name, model_used in methods_info:
    m_data = summary_data.get("methods", {}).get(method_key, {})

    ws_summary.cell(row=row_idx, column=1, value=display_name).font = font_bold_data
    ws_summary.cell(row=row_idx, column=2, value=model_used).alignment = align_left

    # Accuracy
    acc = m_data.get("accuracy", 0)
    ws_summary.cell(row=row_idx, column=3, value=f"{acc:.1f}%").alignment = align_center

    # Precision
    prec = m_data.get("macro_precision", 0)
    ws_summary.cell(row=row_idx, column=4, value=f"{prec:.2f}%").alignment = align_center

    # Recall
    rec = m_data.get("macro_recall", 0)
    ws_summary.cell(row=row_idx, column=5, value=f"{rec:.2f}%").alignment = align_center

    # F1
    f1 = m_data.get("macro_f1_score", 0)
    ws_summary.cell(row=row_idx, column=6, value=f"{f1:.2f}%").alignment = align_center

    # Latency
    lat = m_data.get("average_latency_s", 0)
    ws_summary.cell(row=row_idx, column=7, value=f"{lat:.3f}s").alignment = align_center

    # Hallucination
    halluc = m_data.get("hallucination_rate", 0)
    ws_summary.cell(row=row_idx, column=8, value=f"{halluc:.1f}%").alignment = align_center

    # Highlight ACIS row
    if method_key == "acis":
        for c in range(1, 9):
            ws_summary.cell(row=row_idx, column=c).fill = fill_acis
            ws_summary.cell(row=row_idx, column=c).font = font_bold_data
    else:
        # Zebra striping for others
        if row_idx % 2 == 0:
            for c in range(1, 9):
                ws_summary.cell(row=row_idx, column=c).fill = fill_zebra

    # Borders
    for c in range(1, 9):
        cell = ws_summary.cell(row=row_idx, column=c)
        if cell.font != font_bold_data:
            cell.font = font_data
        cell.border = border_cell

    row_idx += 1

# Add Delta summary rows (Comparison with other baselines)
row_idx += 1
ws_summary.cell(row=row_idx, column=1, value="Cải tiến của hệ thống ACIS:").font = font_section

acis_d = summary_data.get("methods", {}).get("acis", {})
for target_key, target_display in [("chatgpt", "vs ChatGPT (gpt-4o-mini)"), ("gemini", "vs Gemini (gemini-1.5-flash-8b)"), ("grok", "vs Groq (llama-3.1-8b-instant)")]:
    row_idx += 1
    ws_summary.cell(row=row_idx, column=1, value=f"ACIS {target_display}").font = Font(name="Arial", size=9, bold=True, italic=True)
    
    tgt_d = summary_data.get("methods", {}).get(target_key, {})
    delta_acc = acis_d.get("accuracy", 0) - tgt_d.get("accuracy", 0)
    delta_prec = acis_d.get("macro_precision", 0) - tgt_d.get("macro_precision", 0)
    delta_rec = acis_d.get("macro_recall", 0) - tgt_d.get("macro_recall", 0)
    delta_f1 = acis_d.get("macro_f1_score", 0) - tgt_d.get("macro_f1_score", 0)
    delta_halluc = acis_d.get("hallucination_rate", 0) - tgt_d.get("hallucination_rate", 0)
    
    for col_idx, val in [(3, delta_acc), (4, delta_prec), (5, delta_rec), (6, delta_f1)]:
        sign = "+" if val > 0 else ""
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=f"{sign}{val:.2f}%")
        cell.alignment = align_center
        cell.font = Font(name="Arial", size=9, bold=True, color="008000" if val > 0 else "8B0000")
        cell.border = border_cell
        
    # Hallucination delta (negative is better)
    sign = "" if delta_halluc < 0 else "+"
    cell = ws_summary.cell(row=row_idx, column=8, value=f"{sign}{delta_halluc:.2f}%")
    cell.alignment = align_center
    cell.font = Font(name="Arial", size=9, bold=True, color="008000" if delta_halluc <= 0 else "8B0000")
    cell.border = border_cell
    
    # Empty details for others
    ws_summary.cell(row=row_idx, column=2, value="—").alignment = align_left
    ws_summary.cell(row=row_idx, column=2).border = border_cell
    ws_summary.cell(row=row_idx, column=7, value="—").alignment = align_center
    ws_summary.cell(row=row_idx, column=7).border = border_cell

# ==========================================
# SHEET 2: DETAILED RESULTS
# ==========================================
ws_detail = wb.create_sheet(title="Kết quả chi tiết")

ws_detail["A1"] = "CHI TIẾT KẾT QUẢ PHÂN LOẠI TỪNG MẪU ẢNH (ACIS vs 3 AI ĐƠN LẺ)"
ws_detail["A1"].font = font_title

headers_detail = [
    "STT", "Tên File", "Nhãn Gốc (Ground Truth)",
    "Dự đoán ACIS", "ACIS Đúng?", 
    "Dự đoán ChatGPT", "ChatGPT Đúng?", 
    "Dự đoán Gemini", "Gemini Đúng?", 
    "Dự đoán Groq", "Groq Đúng?"
]

for col_idx, text in enumerate(headers_detail, 1):
    cell = ws_detail.cell(row=3, column=col_idx, value=text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

detail_row_idx = 4
for item in detailed_data:
    ws_detail.cell(row=detail_row_idx, column=1, value=item.get("id")).alignment = align_center
    ws_detail.cell(row=detail_row_idx, column=2, value=item.get("filename")).alignment = align_left
    ws_detail.cell(row=detail_row_idx, column=3, value=item.get("ground_truth")).alignment = align_left

    methods_data = item.get("methods", {})

    # ACIS
    acis = methods_data.get("acis", {})
    ws_detail.cell(row=detail_row_idx, column=4, value=acis.get("predicted_label") or "Không xác định").alignment = align_left
    acis_ok = ws_detail.cell(row=detail_row_idx, column=5, value="ĐÚNG" if acis.get("is_correct") else "SAI")
    acis_ok.alignment = align_center
    acis_ok.fill = fill_green if acis.get("is_correct") else fill_red
    acis_ok.font = font_bold_data

    # ChatGPT
    chatgpt = methods_data.get("chatgpt", {})
    ws_detail.cell(row=detail_row_idx, column=6, value=chatgpt.get("predicted_label") or "Không xác định").alignment = align_left
    cg_ok = ws_detail.cell(row=detail_row_idx, column=7, value="ĐÚNG" if chatgpt.get("is_correct") else "SAI")
    cg_ok.alignment = align_center
    cg_ok.fill = fill_green if chatgpt.get("is_correct") else fill_red

    # Gemini
    gemini = methods_data.get("gemini", {})
    ws_detail.cell(row=detail_row_idx, column=8, value=gemini.get("predicted_label") or "Không xác định").alignment = align_left
    gem_ok = ws_detail.cell(row=detail_row_idx, column=9, value="ĐÚNG" if gemini.get("is_correct") else "SAI")
    gem_ok.alignment = align_center
    gem_ok.fill = fill_green if gemini.get("is_correct") else fill_red

    # Groq
    grok = methods_data.get("grok", {})
    ws_detail.cell(row=detail_row_idx, column=10, value=grok.get("predicted_label") or "Không xác định").alignment = align_left
    groq_ok = ws_detail.cell(row=detail_row_idx, column=11, value="ĐÚNG" if grok.get("is_correct") else "SAI")
    groq_ok.alignment = align_center
    groq_ok.fill = fill_green if grok.get("is_correct") else fill_red

    for c in range(1, 12):
        cell = ws_detail.cell(row=detail_row_idx, column=c)
        if cell.font != font_bold_data:
            cell.font = font_data
        cell.border = border_cell

    detail_row_idx += 1

# ==========================================
# AUTO-FIT COLUMN WIDTHS
# ==========================================
for ws in [ws_summary, ws_detail]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ["A1", "A2", "A4"]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save
output_excel = "experiment_results/thuc_nghiem_dataset1_chinh_xac.xlsx"
wb.save(output_excel)
print(f"Excel report saved: {output_excel}")

# Verify
print("\n=== VERIFICATION ===")
for mk, mn, _ in methods_info:
    m = summary_data["methods"][mk]
    print(f"{mn}: Acc={m['accuracy']}%, Prec={m['macro_precision']}%, Rec={m['macro_recall']}%, F1={m['macro_f1_score']}%, Lat={m['average_latency_s']}s, Halluc={m['hallucination_rate']}%")
