"""
Export experiment results to Excel file.
Reads detailed_results.json from both datasets and creates a formatted Excel workbook.
"""
import json
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).parent / "experiment_results"
DATASETS = {
    "dataset1_video_lens5": "Dataset 1 (Video thực tế)",
    "dataset2_ai_lens5": "Dataset 2 (AI tổng hợp)",
}
OUTPUT = Path(__file__).parent / "experiment_results" / "ket_qua_thuc_nghiem.xlsx"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
error_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def load_json(dataset_id):
    path = BASE / dataset_id / "detailed_results.json"
    if not path.exists():
        print(f"  [Skip] {path} not found")
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def add_detail_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(title=sheet_name[:31])
    
    # Headers
    headers = ["#", "Filename", "Ground Truth", "Method", "Raw Prediction", "Predicted Label",
               "Confidence", "Correct?", "Hallucinated?", "Latency (s)", "Error"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    row = 2
    for item in data:
        gt = item.get("ground_truth", "")
        methods = item.get("methods", {})
        fname = item.get("filename", "")
        img_id = item.get("id", "")
        
        for method_name in ["chatgpt", "grok", "gemini", "acis"]:
            m = methods.get(method_name, {})
            raw_pred = m.get("raw_prediction", "")
            pred_label = m.get("predicted_label", "")
            conf = m.get("confidence", 0)
            correct = m.get("is_correct", False)
            halluc = m.get("is_hallucinated", False)
            latency = m.get("latency_s", 0)
            error = m.get("error", "") or ""
            
            values = [img_id, fname, gt, method_name.upper(), raw_pred, pred_label or "",
                      conf, "✓" if correct else "✗", "Yes" if halluc else "No",
                      round(latency, 2), error[:100] if error else ""]
            
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=False)
                
                # Color correct/wrong
                if col == 8:  # Correct column
                    cell.fill = correct_fill if correct else wrong_fill
                if col == 11 and error:  # Error column
                    cell.fill = error_fill
            
            row += 1
    
    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    
    return ws

def add_summary_sheet(wb, all_data):
    ws = wb.create_sheet(title="Tổng hợp", index=0)
    
    headers = ["Dataset", "Method", "Total Images", "Correct", "Accuracy (%)",
               "Hallucinated", "Halluc. Rate (%)", "Avg Latency (s)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    
    row = 2
    for ds_name, items in all_data.items():
        # Compute per-method stats
        method_stats = {}
        for item in items:
            methods = item.get("methods", {})
            for mname in ["chatgpt", "grok", "gemini", "acis"]:
                m = methods.get(mname, {})
                if mname not in method_stats:
                    method_stats[mname] = {"total": 0, "correct": 0, "halluc": 0, "latency": []}
                method_stats[mname]["total"] += 1
                if m.get("is_correct"):
                    method_stats[mname]["correct"] += 1
                if m.get("is_hallucinated"):
                    method_stats[mname]["halluc"] += 1
                method_stats[mname]["latency"].append(m.get("latency_s", 0))
        
        for mname in ["chatgpt", "grok", "gemini", "acis"]:
            s = method_stats.get(mname, {"total": 0, "correct": 0, "halluc": 0, "latency": [0]})
            total = s["total"]
            correct = s["correct"]
            halluc = s["halluc"]
            avg_lat = sum(s["latency"]) / len(s["latency"]) if s["latency"] else 0
            acc = (correct / total * 100) if total > 0 else 0
            halluc_rate = (halluc / total * 100) if total > 0 else 0
            
            display_name = {
                "chatgpt": "ChatGPT (GPT-4o)",
                "grok": "Grok (LLaMA 3.3 70B)",
                "gemini": "Gemini 2.5 Flash",
                "acis": "ACIS (Hệ thống)"
            }.get(mname, mname)
            
            values = [ds_name, display_name, total, correct, round(acc, 2),
                      halluc, round(halluc_rate, 2), round(avg_lat, 2)]
            
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                
                # Highlight ACIS row
                if mname == "acis":
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Separator row
        row += 1
    
    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)
    
    return ws

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    all_data = {}
    for ds_id, ds_label in DATASETS.items():
        print(f"Loading {ds_id}...")
        data = load_json(ds_id)
        if data:
            all_data[ds_label] = data
            add_detail_sheet(wb, ds_label[:31], data)
            print(f"  -> {len(data)} images loaded")
        else:
            print(f"  -> No data")
    
    if all_data:
        add_summary_sheet(wb, all_data)
    
    wb.save(str(OUTPUT))
    print(f"\n[OK] Excel saved: {OUTPUT}")
    print(f"   Sheets: {len(wb.sheetnames)} sheets created")

if __name__ == "__main__":
    main()
