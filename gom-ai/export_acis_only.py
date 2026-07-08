"""Export ACIS-only benchmark results — showing only accuracy metrics."""
import json
import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

results_path = "experiment_results/dataset1_video_lens5/detailed_results.json"
summary_path = "experiment_results/dataset1_video_lens5/summary.json"

with open(results_path, "r", encoding="utf-8") as f:
    detailed_data = json.load(f)
with open(summary_path, "r", encoding="utf-8") as f:
    summary_data = json.load(f)

wb = openpyxl.Workbook()

# Styles
navy = "1B365D"
accent = "E8F1F5"
green_bg = "D4EDDA"
red_bg = "F8D7DA"
white = "FFFFFF"

font_title = Font(name="Arial", size=16, bold=True, color=navy)
font_section = Font(name="Arial", size=12, bold=True, color=navy)
font_header = Font(name="Arial", size=10, bold=True, color=white)
font_data = Font(name="Arial", size=10)
font_bold = Font(name="Arial", size=10, bold=True)
font_big_number = Font(name="Arial", size=28, bold=True, color=navy)

fill_header = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
fill_green = PatternFill(start_color=green_bg, end_color=green_bg, fill_type="solid")
fill_red = PatternFill(start_color=red_bg, end_color=red_bg, fill_type="solid")
fill_accent = PatternFill(start_color=accent, end_color=accent, fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

thin_side = Side(border_style="thin", color="D3D3D3")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# ==========================================
# SHEET 1: ACIS ACCURACY OVERVIEW
# ==========================================
ws = wb.active
ws.title = "Độ chính xác ACIS"

ws["A1"] = "BÁO CÁO ĐỘ CHÍNH XÁC HỆ THỐNG ACIS"
ws["A1"].font = font_title
ws["A2"] = "Tập dữ liệu: Dataset 1 — 100 mẫu ảnh gốm sứ thực tế (trích xuất từ video)"
ws["A2"].font = Font(name="Arial", size=11, italic=True)

# Overall accuracy
acis = summary_data["methods"]["acis"]
overall_acc = acis["accuracy"]

ws["A4"] = "Độ chính xác tổng thể (Overall Accuracy):"
ws["A4"].font = font_section
ws["B4"] = f"{overall_acc:.1f}%"
ws["B4"].font = font_big_number
ws["B4"].alignment = align_center

ws["A5"] = f"Số mẫu đúng: {acis['correct']} / {acis['total']}"
ws["A5"].font = font_data

# Per-class accuracy
ws["A7"] = "Độ chính xác theo từng dòng gốm (Per-class Accuracy):"
ws["A7"].font = font_section

headers = ["Dòng gốm", "Tổng mẫu", "Đúng", "Sai", "Độ chính xác (%)"]
for col_idx, text in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col_idx, value=text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

# Compute per-class accuracy from detailed_results
class_total = Counter()
class_correct = Counter()
for item in detailed_data:
    gt = item["ground_truth"]
    class_total[gt] += 1
    if item["methods"]["acis"].get("is_correct"):
        class_correct[gt] += 1

row = 9
for cls in sorted(class_total.keys()):
    total = class_total[cls]
    correct = class_correct[cls]
    wrong = total - correct
    acc = correct / total * 100 if total > 0 else 0

    ws.cell(row=row, column=1, value=cls).font = font_bold
    ws.cell(row=row, column=2, value=total).alignment = align_center
    ws.cell(row=row, column=3, value=correct).alignment = align_center
    ws.cell(row=row, column=4, value=wrong).alignment = align_center

    acc_cell = ws.cell(row=row, column=5, value=f"{acc:.0f}%")
    acc_cell.alignment = align_center
    acc_cell.font = font_bold

    if acc >= 90:
        acc_cell.fill = fill_green
    elif acc < 50:
        acc_cell.fill = fill_red

    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border_cell
        if ws.cell(row=row, column=c).font == Font():
            ws.cell(row=row, column=c).font = font_data

    row += 1

# ==========================================
# SHEET 2: DETAILED RESULTS (ACIS only)
# ==========================================
ws2 = wb.create_sheet(title="Chi tiết từng mẫu")

ws2["A1"] = "CHI TIẾT KẾT QUẢ PHÂN LOẠI TỪNG MẪU ẢNH (HỆ THỐNG ACIS)"
ws2["A1"].font = font_title

headers2 = ["STT", "Tên File", "Nhãn Gốc", "ACIS Dự đoán", "Kết quả"]
for col_idx, text in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_idx, value=text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

row2 = 4
for item in detailed_data:
    acis_m = item["methods"]["acis"]
    ws2.cell(row=row2, column=1, value=item.get("id")).alignment = align_center
    ws2.cell(row=row2, column=2, value=item.get("filename")).alignment = align_left
    ws2.cell(row=row2, column=3, value=item.get("ground_truth")).alignment = align_left
    ws2.cell(row=row2, column=4, value=acis_m.get("predicted_label") or acis_m.get("raw_prediction") or "—").alignment = align_left

    ok = acis_m.get("is_correct", False)
    result_cell = ws2.cell(row=row2, column=5, value="ĐÚNG" if ok else "SAI")
    result_cell.alignment = align_center
    result_cell.fill = fill_green if ok else fill_red
    result_cell.font = font_bold

    for c in range(1, 6):
        ws2.cell(row=row2, column=c).border = border_cell
        if ws2.cell(row=row2, column=c).font == Font():
            ws2.cell(row=row2, column=c).font = font_data

    row2 += 1

# Auto-fit columns
for ws_sheet in [ws, ws2]:
    for col in ws_sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws_sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

output = "experiment_results/thuc_nghiem_acis_accuracy.xlsx"
wb.save(output)
print(f"Excel saved: {output}")

# Print summary
print(f"\n=== ACIS ACCURACY: {overall_acc}% ({acis['correct']}/{acis['total']}) ===")
for cls in sorted(class_total.keys()):
    t = class_total[cls]
    c = class_correct[cls]
    print(f"  {cls}: {c}/{t} = {c/t*100:.0f}%")
